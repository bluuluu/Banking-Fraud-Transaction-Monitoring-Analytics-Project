"""Generate portfolio artifacts from processed fraud monitoring outputs.

Creates:
    - reports/case_review_summary.xlsx
    - reports/business_insights.md
    - reports/screenshots/*.png
    - notebooks/*.ipynb
"""

from __future__ import annotations

import os
from pathlib import Path

os.environ.setdefault("MPLCONFIGDIR", str(Path(".cache/matplotlib").resolve()))
os.environ.setdefault("XDG_CACHE_HOME", str(Path(".cache").resolve()))
Path(os.environ["MPLCONFIGDIR"]).mkdir(parents=True, exist_ok=True)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import nbformat as nbf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DATA_DIR = Path("data/processed")
REPORTS_DIR = Path("reports")
SCREENSHOTS_DIR = REPORTS_DIR / "screenshots"
NOTEBOOKS_DIR = Path("notebooks")


def load_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    clean = pd.read_csv(DATA_DIR / "clean_transactions.csv", parse_dates=["transaction_ts"])
    features = pd.read_csv(DATA_DIR / "transaction_features.csv", parse_dates=["transaction_ts"])
    scored = pd.read_csv(DATA_DIR / "fraud_scores.csv", parse_dates=["transaction_ts"])
    return clean, features, scored


def save_chart(fig: plt.Figure, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    fig.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def generate_charts(scored: pd.DataFrame) -> None:
    SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)

    risk_counts = scored["risk_band"].value_counts().reindex(["Critical", "High", "Medium", "Low"]).fillna(0)
    fig, ax = plt.subplots(figsize=(8, 5))
    risk_counts.plot(kind="bar", ax=ax, color=["#9b1c31", "#d95f02", "#f0ad4e", "#2b8cbe"])
    ax.set_title("Transactions by Risk Band")
    ax.set_xlabel("Risk Band")
    ax.set_ylabel("Transaction Count")
    save_chart(fig, SCREENSHOTS_DIR / "risk_band_distribution.png")

    merchant_summary = (
        scored.groupby("merchant_category", dropna=False)
        .agg(transaction_count=("transaction_id", "count"), avg_risk_score=("final_risk_score", "mean"))
        .sort_values("avg_risk_score", ascending=False)
        .head(10)
    )
    fig, ax = plt.subplots(figsize=(10, 5))
    merchant_summary["avg_risk_score"].sort_values().plot(kind="barh", ax=ax, color="#4c78a8")
    ax.set_title("Average Risk Score by Merchant Category")
    ax.set_xlabel("Average Final Risk Score")
    ax.set_ylabel("Merchant Category")
    save_chart(fig, SCREENSHOTS_DIR / "merchant_category_risk.png")

    daily = (
        scored.assign(transaction_date=scored["transaction_ts"].dt.date)
        .groupby("transaction_date")
        .agg(
            transactions=("transaction_id", "count"),
            high_risk=("risk_band", lambda s: s.isin(["Critical", "High"]).sum()),
            confirmed_fraud=("fraud_label", "sum"),
        )
        .reset_index()
    )
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(daily["transaction_date"], daily["transactions"], marker="o", label="Transactions")
    ax.plot(daily["transaction_date"], daily["high_risk"], marker="o", label="High Risk")
    ax.plot(daily["transaction_date"], daily["confirmed_fraud"], marker="o", label="Confirmed Fraud")
    ax.set_title("Daily Transaction and Fraud Monitoring Trend")
    ax.set_xlabel("Date")
    ax.set_ylabel("Count")
    ax.tick_params(axis="x", rotation=35)
    ax.legend()
    save_chart(fig, SCREENSHOTS_DIR / "daily_fraud_trend.png")

    top_high_risk = scored.sort_values("final_risk_score", ascending=False).head(10)
    fig, ax = plt.subplots(figsize=(10, 5))
    labels = top_high_risk["customer_name"] + " #" + top_high_risk["transaction_id"].astype(str)
    ax.barh(labels[::-1], top_high_risk["final_risk_score"][::-1], color="#7f3c8d")
    ax.set_title("Top High-Risk Transactions")
    ax.set_xlabel("Final Risk Score")
    ax.set_ylabel("Customer / Transaction")
    save_chart(fig, SCREENSHOTS_DIR / "top_high_risk_transactions.png")


def generate_excel_workbook(clean: pd.DataFrame, features: pd.DataFrame, scored: pd.DataFrame) -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = REPORTS_DIR / "case_review_summary.xlsx"

    summary = pd.DataFrame(
        {
            "Metric": [
                "Total transactions",
                "Total transaction value",
                "Confirmed fraud transactions",
                "High-risk transactions",
                "Average final risk score",
                "Critical transactions",
            ],
            "Value": [
                len(scored),
                round(scored["amount"].sum(), 2),
                int(scored["fraud_label"].sum()),
                int(scored["risk_band"].isin(["Critical", "High"]).sum()),
                round(scored["final_risk_score"].mean(), 2),
                int(scored["risk_band"].eq("Critical").sum()),
            ],
        }
    )
    high_risk = scored[scored["risk_band"].isin(["Critical", "High"])].copy()
    customer_summary = (
        scored.groupby(["customer_id", "customer_name", "customer_segment", "kyc_risk_rating"])
        .agg(
            transaction_count=("transaction_id", "count"),
            total_amount=("amount", "sum"),
            avg_risk_score=("final_risk_score", "mean"),
            high_risk_count=("risk_band", lambda s: s.isin(["Critical", "High"]).sum()),
            confirmed_fraud_count=("fraud_label", "sum"),
        )
        .reset_index()
        .sort_values(["high_risk_count", "avg_risk_score"], ascending=False)
    )
    merchant_summary = (
        scored.groupby(["merchant_category", "merchant_risk_rating"])
        .agg(
            transaction_count=("transaction_id", "count"),
            total_amount=("amount", "sum"),
            avg_risk_score=("final_risk_score", "mean"),
            high_risk_count=("risk_band", lambda s: s.isin(["Critical", "High"]).sum()),
            confirmed_fraud_count=("fraud_label", "sum"),
        )
        .reset_index()
        .sort_values(["high_risk_count", "avg_risk_score"], ascending=False)
    )
    daily_summary = (
        scored.assign(transaction_date=scored["transaction_ts"].dt.date)
        .groupby("transaction_date")
        .agg(
            transaction_count=("transaction_id", "count"),
            total_amount=("amount", "sum"),
            high_risk_count=("risk_band", lambda s: s.isin(["Critical", "High"]).sum()),
            confirmed_fraud_count=("fraud_label", "sum"),
            avg_risk_score=("final_risk_score", "mean"),
        )
        .reset_index()
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Executive Summary", index=False)
        high_risk.to_excel(writer, sheet_name="High Risk Queue", index=False)
        customer_summary.to_excel(writer, sheet_name="Customer Risk", index=False)
        merchant_summary.to_excel(writer, sheet_name="Merchant Risk", index=False)
        daily_summary.to_excel(writer, sheet_name="Daily Trends", index=False)
        features.head(100).to_excel(writer, sheet_name="Feature Sample", index=False)
        clean.head(100).to_excel(writer, sheet_name="Clean Data Sample", index=False)

    workbook = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)
    workbook.save(output_path)


def generate_insights(scored: pd.DataFrame) -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    high_risk = scored[scored["risk_band"].isin(["Critical", "High"])]
    top_customer = (
        scored.groupby("customer_name")["final_risk_score"].mean().sort_values(ascending=False).head(1)
    )
    top_category = (
        scored.groupby("merchant_category")["final_risk_score"].mean().sort_values(ascending=False).head(1)
    )
    confirmed_fraud = int(scored["fraud_label"].sum())
    fraud_rate = confirmed_fraud / len(scored) if len(scored) else 0

    lines = [
        "# Fraud Monitoring Business Insights",
        "",
        "## Executive Summary",
        "",
        f"- Reviewed {len(scored):,} transactions with ${scored['amount'].sum():,.2f} in total transaction value.",
        f"- Flagged {len(high_risk):,} high-risk transactions for analyst review.",
        f"- Identified {confirmed_fraud:,} confirmed fraud-labelled transactions, equal to a {fraud_rate:.1%} fraud rate in the synthetic sample.",
        f"- Highest average customer risk: {top_customer.index[0]} with an average score of {top_customer.iloc[0]:.2f}.",
        f"- Highest average merchant category risk: {top_category.index[0]} with an average score of {top_category.iloc[0]:.2f}.",
        "",
        "## Key Observations",
        "",
        "- International card-not-present transactions drive a large share of high-risk alerts.",
        "- High-risk merchant categories such as Crypto Exchange, Gaming, Gift Cards, and Luxury Goods produce elevated risk scores.",
        "- Velocity, new merchant behavior, and sudden spending increases are useful explainable fraud indicators.",
        "- The Isolation Forest score helps surface unusual transactions, while the rule-based score keeps analyst explanations clear.",
        "",
        "## Recommended Monitoring Actions",
        "",
        "- Prioritize Critical and High risk bands for same-day review.",
        "- Monitor customers with repeated cross-border card-not-present transactions.",
        "- Add stricter controls for high-risk merchant categories and new-device transactions.",
        "- Track false positives once analyst outcomes are available to tune rule thresholds.",
        "",
        "## Generated Visuals",
        "",
        "- `reports/screenshots/risk_band_distribution.png`",
        "- `reports/screenshots/merchant_category_risk.png`",
        "- `reports/screenshots/daily_fraud_trend.png`",
        "- `reports/screenshots/top_high_risk_transactions.png`",
        "",
    ]
    (REPORTS_DIR / "business_insights.md").write_text("\n".join(lines), encoding="utf-8")


def code_cell(source: str):
    return nbf.v4.new_code_cell(source.strip())


def markdown_cell(source: str):
    return nbf.v4.new_markdown_cell(source.strip())


def write_notebook(path: Path, title: str, cells: list) -> None:
    notebook = nbf.v4.new_notebook()
    notebook["cells"] = [markdown_cell(f"# {title}")] + cells
    notebook["metadata"] = {
        "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
        "language_info": {"name": "python", "pygments_lexer": "ipython3"},
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    nbf.write(notebook, path)


def generate_notebooks() -> None:
    NOTEBOOKS_DIR.mkdir(parents=True, exist_ok=True)

    write_notebook(
        NOTEBOOKS_DIR / "01_data_profiling.ipynb",
        "Data Profiling",
        [
            markdown_cell("Load the cleaned transaction dataset and review core data quality indicators."),
            code_cell(
                """
                import pandas as pd

                clean = pd.read_csv("../data/processed/clean_transactions.csv", parse_dates=["transaction_ts"])
                clean.head()
                """
            ),
            code_cell(
                """
                clean.info()
                """
            ),
            code_cell(
                """
                clean[["amount", "current_balance", "days_since_onboarding"]].describe()
                """
            ),
            code_cell(
                """
                clean.groupby(["transaction_channel", "transaction_status"]).size().reset_index(name="transaction_count")
                """
            ),
            code_cell(
                """
                clean.isna().sum().sort_values(ascending=False).head(15)
                """
            ),
        ],
    )

    write_notebook(
        NOTEBOOKS_DIR / "02_feature_engineering.ipynb",
        "Feature Engineering",
        [
            markdown_cell("Review fraud monitoring features created by the Python feature engineering script."),
            code_cell(
                """
                import pandas as pd

                features = pd.read_csv("../data/processed/transaction_features.csv", parse_dates=["transaction_ts"])
                features.head()
                """
            ),
            code_cell(
                """
                flag_cols = [col for col in features.columns if col.endswith("_flag")]
                features[flag_cols].sum().sort_values(ascending=False)
                """
            ),
            code_cell(
                """
                features.groupby("merchant_category").agg(
                    transactions=("transaction_id", "count"),
                    avg_triggered_rules=("triggered_rule_count", "mean"),
                    avg_amount=("amount", "mean")
                ).sort_values("avg_triggered_rules", ascending=False)
                """
            ),
            code_cell(
                """
                features.sort_values("triggered_rule_count", ascending=False)[
                    ["transaction_id", "customer_name", "amount", "merchant_category", "triggered_rule_count"]
                ].head(10)
                """
            ),
        ],
    )

    write_notebook(
        NOTEBOOKS_DIR / "03_fraud_scoring.ipynb",
        "Fraud Scoring",
        [
            markdown_cell("Review final risk scores, risk bands, and high-risk analyst queue outputs."),
            code_cell(
                """
                import pandas as pd

                scored = pd.read_csv("../data/processed/fraud_scores.csv", parse_dates=["transaction_ts"])
                high_risk = pd.read_csv("../data/processed/high_risk_transactions.csv", parse_dates=["transaction_ts"])
                scored.head()
                """
            ),
            code_cell(
                """
                scored["risk_band"].value_counts()
                """
            ),
            code_cell(
                """
                scored.groupby("merchant_category").agg(
                    transactions=("transaction_id", "count"),
                    avg_risk_score=("final_risk_score", "mean"),
                    confirmed_fraud=("fraud_label", "sum")
                ).sort_values("avg_risk_score", ascending=False)
                """
            ),
            code_cell(
                """
                high_risk[[
                    "transaction_id", "customer_name", "amount", "merchant_category",
                    "rule_based_risk_score", "isolation_forest_score",
                    "final_risk_score", "risk_band", "recommended_action"
                ]]
                """
            ),
        ],
    )


def main() -> None:
    clean, features, scored = load_data()
    generate_charts(scored)
    generate_excel_workbook(clean, features, scored)
    generate_insights(scored)
    generate_notebooks()
    print("Generated notebooks, report workbook, charts, and business insights.")


if __name__ == "__main__":
    main()
